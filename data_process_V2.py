import pandas as pd
import sys
import codecs
import numpy as np
import os
import torch
sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
print("中文")
data_path = '/disk1/xly/un-planned_reoperation/data/data_lite.csv'
from transformers import AutoModel, AutoTokenizer

from openpyxl import load_workbook


os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
model_dir = os.path.expanduser('/disk1/xly/un-planned_reoperation/Chinses_bert') 
#print(model_dir)
tokenizer = AutoTokenizer.from_pretrained(model_dir)
print("token ok")
model = AutoModel.from_pretrained(model_dir)
print("model ok")



# 加载 Excel 文件
file_path = "/disk1/xly/un-planned_reoperation/data/2025_12_15/extenerel.xlsx"
workbook = load_workbook(file_path)

# 选择工作表
sheet = workbook.active  # 或者 workbook["工作表名称"]
print(sheet)
# 遍历读取数据
j = 0
k = 0
data_encode = []
text_encode = []
for row in sheet.iter_rows(values_only=True):
    row_encode = []

    j = j+1
    if j == 1:
        continue
    # if j ==204:###last
        # break
    
    else:
        print(j)
        #print(row)  # 每行数据会以元组形式输出
        for i in range(len(row)):
            #print(row[i])

            if i == 0:
               row_encode.append(int(row[i]))
            if i == 1:
               if "出血" in row[i]:
                   row_encode.append(1)
               elif "感染" in row[i]:
                   row_encode.append(2)
               else:
                   row_encode.append(0)

            if i == 2:
               if row[i] == None:
                   row_encode.append(-1)
               elif "男" in row[i]:
                   row_encode.append(1)
               elif "女" in row[i]:
                   row_encode.append(0) 
               else:
                   row_encode.append(-1)
                   
            if i == 3:
               row_encode.append(row[i])

            if i == 4:
               print(row[i])
               if row[i] == None:
                   row_encode.append(-1)
               elif row[i] == "I":
                  row_encode.append(1)
               elif row[i] == "II":
                  row_encode.append(2)               
               elif row[i] == "III":
                  row_encode.append(3)
               elif row[i] == "IV":
                  row_encode.append(4)  
               elif row[i] == "V":
                  row_encode.append(5)
               elif row[i] == "IE":
                  row_encode.append(6)
               elif row[i] == "IIE":
                  row_encode.append(7)
               elif row[i] == "IIIE":
                  row_encode.append(8)
               elif row[i] == "EIII":
                  row_encode.append(8)
               elif row[i] == "IVE":
                  row_encode.append(9)
               elif row[i] == "VE":
                  row_encode.append(9)    
               else:
                  row_encode.append(-1)
                      
               # elif row[i] == None:
                  # row_encode.append(-1)
               # elif row[i] == 0:
                  # row_encode.append(0)     
                  
            if i == 5:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)
    
 
            if i == 6:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


 
            if i == 7:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)
  
            if i == 9:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)

               


            if i == 10:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 11:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 12:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 13:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)



            if i == 14:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 15:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 16:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 17:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)

 

            if i == 18:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            if i == 19:
               if row[i] == None:
                   row_encode.append(-1)
                   continue
               elif "是" in row[i]:
                   row_encode.append(1)
               elif "否" in row[i]:
                   row_encode.append(0)
               else:
                   row_encode.append(-1)


            # if i == 20:
               # if row[i] == None:
                   # row_encode.append(2)
                   # continue
               # if "是" in row[i]:
                   # row_encode.append(1)
               # if "否" in row[i]:
                   # row_encode.append(0)  
                   
            if i == 20:
                if type(row[i]) == int:
                    row_encode.append(row[i])  
                elif type(row[i]) == float:
                   row_encode.append(row[i])  
                else:
                    row_encode.append(-1)  
                
            if i == 21:
                if type(row[i]) == int:
                   row_encode.append(row[i]) 
                elif type(row[i]) == float:
                   row_encode.append(row[i])                  
                else:
                    row_encode.append(-1)  
                
            if i == 22:
                if type(row[i]) == float:
                   row_encode.append(row[i])  
                else:
                    row_encode.append(-1)    
            if i== 23:
                # print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(0,v.shape)                    
 
            if i== 24:
                # print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v1 = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(1,v1.shape)

            if i== 25:
                # print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v2 = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(2,v2.shape)
                
                
            if i== 26:
                # print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v3 = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(3,v3.shape)                
                
                
            if i== 27:
                # print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v4 = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(4,v4.shape)                 
                
            if i== 28:
                # print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v5 = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(5,v5.shape)                 
                
                
            if i== 29:
                print(row[i])
                inputs = tokenizer(row[i], return_tensors='pt')
                outputs = model(**inputs)  # shape (1, 7, 768)
                #print(outputs)
                v6 = torch.mean(outputs[0], dim=1)  # shape (1, 768)
                #print(v)
                print(v6.shape)                 
                v = torch.cat([v,v1,v2, v3, v4, v5, v6],0).view(-1).detach().numpy()
                print(6,v.shape)
                text_encode.append(v)
            
        print(row_encode,len(row_encode))        
        if len(row_encode) != 22:     
            print(row,"AAAAAAAAAAA")
            print(row_encode,len(row_encode))
            k=k+1
        
        # print(row)
        # print(row_encode)
        data_encode.append(np.array(row_encode))
print(k)
data_encode = np.array(data_encode)
text_encode = np.array(text_encode)
print(data_encode.shape)
print(data_encode[:,:6])
print(text_encode.shape)
np.savetxt("/disk1/xly/un-planned_reoperation/data/2025_12_15/data_Lite_externel_3.txt", data_encode)
np.savetxt("/disk1/xly/un-planned_reoperation/data/2025_12_15/data_Lite_externel_text_CB.txt", text_encode)



     
